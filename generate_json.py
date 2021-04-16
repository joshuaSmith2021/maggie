import json

from openpyxl import load_workbook

wb = load_workbook(filename='Openings.xlsx')
openings = [x for x in wb.sheetnames if x != 'Template']

result = []

for opening in openings:
    lines = list(wb[opening].rows)
    result += [x[1].value for x in lines[1:]]

with open('Openings.json', 'w') as f:
    f.write(json.dumps(result))
